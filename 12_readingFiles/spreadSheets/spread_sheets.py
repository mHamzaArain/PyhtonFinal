import openpyxl
import os 

os.chdir(r'C:\Users\Family\Desktop\Practice_v1.7\12_readingFiles\spreadSheets')

workbook = openpyxl.load_workbook('example.xlsx')
workbook.get_sheet_names()

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))      # <class 'openpyxl.worksheet.worksheet.Worksheet'>

cell = sheet['A1']
cell.value
print(str(cell.value))      #2015-04-05 13:34:02

print(sheet['B1'].value)        # Apples

print(sheet['C1'].value)        # 73

print(sheet.cell(row=1, column=2))      # <Cell 'Sheet1'.B1>

print(sheet['B1'])     # <Cell 'Sheet1'.B1>

for i in range(1, 8):
	print(i, sheet.cell(row=i, column=2).value)
# 1 Apples
# 2 Cherries
# 3 Pears
# 4 Oranges
# 5 Apples
# 6 Bananas
# 7 Strawberries

print(sheet.max_row)			# Total row (1-7)
print(sheet.max_column)		# Total column (1-3)

# print(openpyxl.cell.get_column_letter(1))
print()
print()
print()
print()
print(help(openpyxl.cell))