# To handle spread sheets
import openpyxl

# var = openpyxl.load_workbook(.xlsx)                         # Load .xlsx file
# workbook.get_sheet_names()                                    # ['Sheet1', 'Sheet2']
# sheet = workbook.get_sheet_by_name('Sheet1')      # get specific sheet
# sheet['B1'].value                                                         # get value 
# sheet.cell(row=i, column=2).value                              # sheet['B1'].value        
# sheet.max_row			                                                    # Total row (1-7)
# sheet.max_column		                                                 # Total column (1-3)



# for i in range(1, 8):                                                         # Printing values in row
# 	print(i, sheet.cell(row=i, column=2).value)